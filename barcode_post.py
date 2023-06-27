from openpyxl.workbook import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Color, PatternFill, Font, Border
# from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import numpy as np
import os
import cv2

def post_processing(vals, labels, filesNames,  temp_imgs, save_path):
    
    if len(vals) > 0:
        wb = Workbook()
        ws = wb.active
        ws.title = "new table"        
        pre_rows = 1
        ws.cell(row=pre_rows, column=1).value = "FILE_NAME"
        ws.cell(row=pre_rows, column=1).font = Font(bold=True)
        ws.cell(row=pre_rows, column=2).value = "QRCODE"
        ws.cell(row=pre_rows, column=2).font = Font(bold=True)
        ws.cell(row=pre_rows, column=3).value = "TOTAL_VALUE"
        ws.cell(row=pre_rows, column=3).font = Font(bold=True)
        ws.cell(row=pre_rows, column=4).value = "TOTAL_IMAGE"
        ws.cell(row=pre_rows, column=4).font = Font(bold=True)
        ws.cell(row=pre_rows, column=5).value = "CROP_PATH"
        ws.cell(row=pre_rows, column=5).font = Font(bold=True)                        
        # ws.cell(row=pre_rows, column=2).value = "BARCODE_1"
        # ws.cell(row=pre_rows, column=2).font = Font(bold=True)
        # ws.cell(row=pre_rows, column=3).value = "BARCODE_2"
        # ws.cell(row=pre_rows, column=3).font = Font(bold=True)
        # ws.cell(row=pre_rows, column=4).value = "QRCODE_1"
        # ws.cell(row=pre_rows, column=4).font = Font(bold=True)
        # ws.cell(row=pre_rows, column=5).value = "QRCODE_2"
        # ws.cell(row=pre_rows, column=5).font = Font(bold=True)
        # ws.cell(row=pre_rows, column=6).value = "TOTAL_VALUE"
        # ws.cell(row=pre_rows, column=6).font = Font(bold=True)
        # ws.cell(row=pre_rows, column=7).value = "TOTAL_IMAGE"
        # ws.cell(row=pre_rows, column=7).font = Font(bold=True)

        pre_rows = pre_rows + 1
        width = [40, 19, 10, 30, 40]
        img_h, img_w = 25, 30
        thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')) 
        redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

        ### loopping every document in multi document ###
        ref_prob = 0.7
        cnt = 0
        total_img_name = os.listdir(temp_imgs)
        total_img_name.sort()
        uiCellW, uiCellH = 130, 50
        uiWCnt = 50
        uiHCnt = int(np.ceil(len(vals)/uiWCnt))
        croppedImgs = np.zeros((uiHCnt * uiCellH, uiWCnt * uiCellW), np.uint8)
        pH, pW = 0, 0
        for k, v in enumerate(vals):
            try:
                # QR1, QR2, Bar1, Bar2 = v[0]
                QR = v[0]
                cnt = cnt + 1
                # cv2.imwrite(total_img_name, total_img)
                filename = filesNames[k]
                ws.cell(row=pre_rows, column=1).value = filename
                ws.cell(row=pre_rows, column=2).value = QR
                # try: ws.cell(row=pre_rows, column=2).value = Bar1
                # except: ws.cell(row=pre_rows, column=2).value = "xxxxxxxx"
                # try: ws.cell(row=pre_rows, column=3).value = Bar2
                # except: ws.cell(row=pre_rows, column=3).value = "xxxxxxxx"
                # ws.cell(row=pre_rows, column=3).value = 'Not_Reconized'
                # try: ws.cell(row=pre_rows, column=4).value = QR1
                # except: ws.cell(row=pre_rows, column=4).value = "xxxxxxxx"
                # ws.cell(row=pre_rows, column=4).value = 'Not_Reconized'
                # try: ws.cell(row=pre_rows, column=5).value = QR2
                # except: ws.cell(row=pre_rows, column=5).value = "xxxxxxxx"
                ws.cell(row=pre_rows, column=3).value = labels[k][0]
                ws.cell(row=pre_rows, column=5).value = temp_imgs + '/' + total_img_name[k]
                img = Image(os.path.join(temp_imgs,total_img_name[k]))
                im = cv2.imread(os.path.join(temp_imgs,total_img_name[k]), 0)
                im = cv2.resize(im,(uiCellW, uiCellH), interpolation = cv2.INTER_CUBIC)
                croppedImgs[pH:pH+uiCellH, pW:pW+uiCellW] = im
                pW = pW+uiCellW
                if pW >= uiWCnt * uiCellW:
                    pW, pH = 0, pH+uiCellH

                ###########################
                col = 3
                row_cut = 9500
                col_cut = 15000
                
                _from = AnchorMarker(
                    col=col,
                    row=pre_rows-1,
                    colOff=col_cut,
                    rowOff=row_cut,
                )
                to = AnchorMarker(
                    col=col + 1,
                    row=pre_rows,
                    colOff=-col_cut,
                    rowOff=-row_cut,
                )
                img.anchor = TwoCellAnchor(editAs="twoCell", _from=_from, to=to)
                ws.add_image(img)
                # if labels[k][1] < ref_prob: ws[get_column_letter(6)+str(pre_rows)].fill = redFill
                pre_rows = pre_rows+1   
            except:
                print(k, "---------------------")   
        # save cropped images:
        # cv2.imwrite(os.path.join(temp_imgs[:-4], 'croppedImgs.jpg'), croppedImgs)
        # cell swrap, thin
        for i, row in enumerate(ws.rows):
            for j in range(len(row)):
                ws[get_column_letter(j+1)+str(i+1)].alignment = Alignment(wrap_text=True, vertical='center',horizontal='center')
            ws.row_dimensions[i+1].height = img_h
            # row_no = row_no + 1
        for i in range(1, pre_rows):       
            for j in range(1, 6):
                ws.cell(row=i, column=j).border = thin_border
        for i, wid in enumerate(width):
            ws.column_dimensions[get_column_letter(i+1)].width = wid
        # save
        wb.save(save_path)
        for cn in range(cnt):
            try:
                os.remove(f"config/temp{cn}.jpg")
            except:
                pass              
    else:
        print("=== Table of this pdf is not detected ===")

    return None
