from PyQt5 import QtWidgets, uic,QtGui, QtCore
from PyQt5.QtWidgets import QMainWindow, QApplication, QVBoxLayout, QDialog
from PyQt5.QtCore import pyqtSignal, pyqtSlot
import sys, os, getpass, shutil
from threading import Thread
import uuid
from AESCipher import AESCipher
import requests, json, os
import openpyxl
import pandas
from main_barcode import barcode
import ctypes

#uuid.getnode()
class MainPage(QMainWindow):
    def __init__(self):
        super(MainPage,self).__init__()     
        uic.loadUi(resource_path("main.ui"),self)
        self.cnt2=0
        self.totol2=0
        self.cnt1=0
        self.totol1=0
        self.initUI() 
    
    def initUI(self):

        self.btn_browse2.clicked.connect(self.btnSelectFile2)
        self.btn_process2.clicked.connect(self.btnProcessFile2)
        self.pbar_tab2.valueChanged.connect(self.onChangeValue2)
        self.lnk_tab2.clicked.connect(self.onClickLblTab2)
        self.single_done2.connect(self.progress2)
        
        # winScale = ctypes.windll.shcore.GetScaleFactorForDevice(0) / 100 # get windows scale
        
        # self.btn_process2.setEnabled(False)

    def openOutput(self,path): 
        if path:  
            QtGui.QDesktopServices.openUrl(QtCore.QUrl.fromLocalFile(str(path)))
    
    def btnSelectFile2(self):
        self.configPath, _ = QtWidgets.QFileDialog.getOpenFileNames(
            self, "Open File", "", "CSV Files(*.json);;All Files (*)",
            options=QtWidgets.QFileDialog.DontUseNativeDialog)
        try:
            self.configPath = self.configPath[0]
            self.le_tab2.setText(self.configPath)   
        except: pass
        print("okay")
    def btnProcessFile2(self):
        t1=Thread(target=self.Operation2)
        t1.start()
        # t1.join()

    def process(self, out, img_list, cropPathBody, batchName, ind, sourcePath):
        self.lbl1_tab2.setText(f"Batch Name: {batchName} ==> scan starting ...\nIf you don't have gpu, startup time will be longer ... ")        
        makedir(out)
        save_path = out + f"/barcode_{batchName}.xlsx"        
        vals, labels, success_imgs = barcode(self,img_list, out, cropPathBody, save_path) # 1 means file name, and 0 means qrcode
        tempSourcePath = '/'.join(sourcePath.split('/')[0:-1] + ['temp.xlsx'])
        if vals:
            try: shutil.copyfile(sourcePath, tempSourcePath)
            except: pass
            wb = openpyxl.load_workbook(tempSourcePath)
            ws = wb.active
            # json output save
            jsonOut = {}
            for i, val in enumerate(vals):
                value = {}
                value["BATCHNAME"] = batchName
                value["QRCODE"] = val[0]
                value["TOTAL VALUE"] = labels[i][0]
                value["CROP IMAGE PATH"] = cropPathBody.replace("\\", "/") + f'/{val[1]}'
                value["ORIGINAL IMG PATH"] = success_imgs[i]
                jsonOut[success_imgs[i]] = value
            with open(os.path.join(out, f"barcode_{batchName}.json"), 'w') as f:
                json.dump(jsonOut, f, indent=4, sort_keys=True)  
            for i in ind:
                ws.cell(i+2,3).value = 'Y' 
            
            wb.save(sourcePath)
            try: os.remove(tempSourcePath)
            except: pass   
            for imd in img_list:
                try: shutil.copytree(imd, os.path.join(out, 'inputs', imd.split('\\')[-1]))
                except: pass                     
            return True              
        else:
            return False   

        return None    

    def Operation2(self):
        try:
            self.configPath = self.le_tab2.text()
            with open(self.configPath, "r") as f:
                jsonInfos = json.load(f)
            sourcePath = jsonInfos['excelFilePath']
            self.outputPath = jsonInfos['outputPath']
            makedir(self.outputPath)
        except: 
            self.lbl1_tab2.setText("Wanring: Please Select Config File Exactly")
            return None        
        # wb_obj = openpyxl.load_workbook(sourcePath)
        # sheet_obj = wb_obj.active
        source = pandas.read_excel(sourcePath).to_numpy()
        batchName, cropPathBody = None, "CROP"
        img_list = []
        
        ind = []
        for i, item in enumerate(source):
            if str(item[2]) == 'None' or str(item[2]) == 'nan': break
            if item[2].lower() == 'n':
                if batchName != item[0]:
                    try: 
                        self.process(out, img_list, os.path.join(out, cropPathBody) , batchName, ind, sourcePath)
                        img_list = []
                    except: pass
                    batchName = item[0]
                    out = os.path.join(self.outputPath, batchName)
                img_list.append(item[1])
                ind.append(i)
        if len(img_list) > 0:
            self.process(out, img_list, os.path.join(out, cropPathBody), batchName, ind, sourcePath)        
        self.lnk_tab2.setEnabled(True)
        self.lnk_tab2.setText("Go Output Folder")
        self.lbl1_tab2.setText("Done!") 

    def onChangeValue2(self,val):
        self.pbar_tab2.setFormat(str(self.cnt2) + '/' + str(self.total2))

    single_done2 = pyqtSignal()
    @pyqtSlot()
    def progress2(self):
        self.pbar_tab2.setValue(int((self.cnt2/self.total2)*100))

    def openFolder2(self, path):
        self.lbl1_tab2.setText("Result : Successfully processed ")
        self.path2=path
        self.lnk_tab2.setText(str(path))
        self.openOutput(path)
    
    def onClickLblTab2(self):
        self.openFolder2(self.outputPath)
        # self.openOutput(self.path2)
        #QtGui.QDesktopServices.openUrl(QtCore.QUrl.fromLocalFile(str(path)))

def window():
    app = QApplication(sys.argv)
    win = MainPage()
    win.show()
    sys.exit(app.exec_())

def windowValidate():
    app = QApplication(sys.argv)
    win = KeyWindow()
    win.show()
    sys.exit(app.exec_())


class KeyWindow(QMainWindow):
    def __init__(self):
        super(KeyWindow,self).__init__()
        self.node = str(uuid.getnode())
        uic.loadUi(resource_path("keyWindow.ui"),self)
        self.lbl_id.setText(self.node)
        self.initUI()

    def initUI(self):
        self.btn_submit.clicked.connect(self.onSubmit)
    
    def onSubmit(self):
        
        if self.txt_key.toPlainText():
            fp = open(logPath+'/.validate', 'w')
            fp.write(self.txt_key.toPlainText())
            fp.close()

            if validate():
                self.win = MainPage()
                self.win.show()
                self.hide()
            else:
                self.lbl_msg.setText("Invalid key")
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def validate():
    try:
        fp =open(logPath+'/.validate','rb')
        data = fp.read()
        fp.close()
    except: return False
    if data:
        try:
            c = AESCipher()
            if str(uuid.getnode()) == c.decrypt(data):
                return True
        except:
            pass
    return False
def makedir(dir):
    try:
        os.mkdir(dir)
    except:
        pass

logPath = f"C:/Users/{getpass.getuser()}/.barcode"
makedir(logPath)  
try: shutil.copytree("config", os.path.join(logPath, 'config'))
except: pass

severURL = 'http://127.0.0.1:8001'
# window()

if not validate():
    windowValidate()
else:
    window()

#window()